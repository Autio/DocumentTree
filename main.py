# 4/7/2017 Document tree creator for Newdea Project Center export - P. Autio
# Newdea Project Center allows you to export all the files it contains as an archive file
# There all the files have a physical filename which is a GUID
# A table in the Project Center export database contains a mapping of what these files are called
# in the system. It is possible to do lookups against the hierarchy tables in this PC export
# to also establish where in the system these files are

# The aim of tihs code is to take such a lookup and copy the PC export attachment files from a flat structure
# and obsucre filenames into the proper filenames in a folder structure that matches the structure in PC which
# for WaterAid is: Region -> CP -> Programme -> Project -> Sub-project

# Pseudocode:
# 1. Open file that contains the hierarchy and name mapping
# 2. Read in the filepaths and filenames
# 3. Got through and copy/move the files whilst renaming them
from openpyxl import load_workbook, Workbook

import os
import shutil

filename = "FixedHierarchy.xlsx"
filepath = ""

wb = load_workbook(filename, data_only = True)

sheet = wb['FixedHierarchy']
row_count = sheet.max_row
files = [["","",""] for y in range(row_count-1)] # Headers not needed



def copy_rename(newFileFolder, oldFileName, newFileName):
    # assume the source is in the current directory
    srcDir = os.curdir
    dstDir = os.path.join(os.curdir, newFileFolder)
    srcFile = os.path.join(srcDir, oldFileName)

    shutil.copy(srcFile, dstDir)

    dstFile = os.path.join(dstDir, oldFileName)
    newDstFileName = os.path.join(dstDir, newFileName)
    os.rename(dstFile, newDstFileName)


r = 2
while r <= row_count:

    region, CP, programme, project, subProject, fileName, rawName = "", "", "", "", "", "", ""

    try:
        region = str(sheet['B' + str(r)].value)
    except:
        print("Could not get region")
    try:
        CP = str(sheet['C' + str(r)].value)
    except:
        print("Error in getting CP")
    try:
        programme = str(sheet['D'+ str(r)].value)
    except:
        print("Error in getting programme")
    try:
        project = str(sheet['E' + str(r)].value)
    except:
        print("Error in getting project")
    try:
        subProject = str(sheet['F' + str(r)].value)
    except:
        print("Error in getting subproject")
    try:
        fileName = str(sheet['J' + str(r)].value)
    except:
        print("Error in getting filename")
    try:
        rawName = str(sheet['K' + str(r)].value)
    except:
        print("Error in getting rawname")

    # Concatenate folder structure
    if(region == "#N/A"):
        # don't output file, there's something wrong with it
        filepath = "#N/A"
    else:
        filepath = region + "\\"
        if (CP != "#N/A" and CP != "None"):
            filepath += CP + "\\"
            if (programme != "#N/A" and programme != "None"):
                filepath +=  programme + "\\"
                if (project != "#N/A" and project != "None"):
                    filepath += project + "\\"
                    if (subProject != "#N/A"and subProject != "None"):
                        filepath += subProject + "\\"

    files[r - 2][0] = filepath
    files[r - 2][1] = rawName
    files[r - 2][2] = fileName

    r += 1


# Go through created list and move and rename files
i = 0
while i < len(files):
    targetFilePath = files[i][0].replace(':','')
    # Check if folder exists, and if it doesn't then create it
    if not (targetFilePath == "#N/A"):
        if not (os.path.isdir(targetFilePath)):
            os.makedirs(targetFilePath)
        # copy and rename
        # assume the source is in the current directory
        copy_rename(targetFilePath, files[i][1], files[i][2])


    i += 1
